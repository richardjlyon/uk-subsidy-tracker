"""Ofgem aggregate-grain loaders (Phase 05.2).

Raw artefacts consumed:
  - data/raw/ofgem/rocs_report_2006_to_2018_20250410081520.xlsx  (12-year XLSX primary)
  - data/raw/ofgem/ro-annual-aggregate.csv                        (SY17–SY23 transcription)
  - data/raw/ofgem/roc-prices.csv                                 (buyout + recycle + eroc + mutualisation)
  - data/raw/ofgem/ro-generation.csv                              (regenerated from XLSX, D-01)

Provenance: Ofgem 12-year XLSX is the authoritative pre-2019 source (GB aggregate
by technology x month x country). Post-2018 aggregates are transcribed from
Ofgem annual-report PDFs — see each transcribed CSV's header Provenance block
and `.meta.json::sources[]` array for per-row source URL + sha256 + retrieved_on
(D-03 multi-source sidecar discipline).

Error-path discipline (Phase 4 D-17 + Plan 07):
  - `output_path` bound BEFORE `try:` block
  - `timeout=60` on `requests.get()`
  - bare `raise` in `except requests.exceptions.RequestException`
  - successful download returns `output_path`

Loader-owned pandera validation per Phase 2 D-04: each `load_*_csv()` validates
its DataFrame against a module-level schema before returning.

Wave scoping (Phase 05.2 Plan 01):
  - This module lands the loader skeleton + download function + pandera schemas.
  - `parse_xlsx_to_monthly()` is implemented in Plan 02 (replaces the Plan 01 stub).
    Reads ``ROCs by Tech & Month`` sheet from the 12-year XLSX and emits monthly
    ROC counts by technology. See also ``emit_ro_generation_csv()``.
"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pandera.pandas as pa
import requests
import yaml
from pydantic import BaseModel

from uk_subsidy_tracker import DATA_DIR
from uk_subsidy_tracker.data.utils import HEADERS


# ---------------------------------------------------------------------------
# Module constants — single source of truth for the 12-year XLSX upstream.
# ---------------------------------------------------------------------------
XLSX_URL = (
    "https://www.ofgem.gov.uk/sites/default/files/2025-05/"
    "rocs_report_2006_to_2018_20250410081520.xlsx"
)
XLSX_FILENAME = "rocs_report_2006_to_2018_20250410081520.xlsx"


# ---------------------------------------------------------------------------
# Pandera schemas — loader-owned validation per Phase 2 D-04.
# ``strict=False`` permits additional un-validated columns (CSVs may carry
# transcribed-source-PDF URLs etc. without breaking schema evolution).
# ---------------------------------------------------------------------------
ofgem_annual_aggregate_schema = pa.DataFrameSchema(
    {
        "scheme_year": pa.Column(str, checks=pa.Check.str_matches(r"^SY\d{2}$")),
        "year": pa.Column("int64", coerce=True),
        "country": pa.Column(str, checks=pa.Check.isin(["GB", "NI"])),
        "technology": pa.Column(str, coerce=True),
        "generation_gwh": pa.Column(float, nullable=True, coerce=True),
        "rocs_issued": pa.Column(float, nullable=True, coerce=True),
        "ro_cost_gbp_nominal": pa.Column(float, nullable=True, coerce=True),
        "source_pdf_url": pa.Column(str, nullable=True, coerce=True),
    },
    strict=False,
    coerce=True,
)

ofgem_roc_prices_schema = pa.DataFrameSchema(
    {
        "obligation_year": pa.Column(
            str,
            checks=pa.Check.str_matches(r"^\d{4}-\d{2}$|^SY\d{2}$"),
        ),
        "buyout_gbp_per_roc": pa.Column(float, coerce=True),
        "recycle_gbp_per_roc": pa.Column(float, nullable=True, coerce=True),
        "eroc_gbp_per_roc": pa.Column(float, nullable=True, coerce=True),
        "mutualisation_gbp_total": pa.Column(float, nullable=True, coerce=True),
    },
    strict=False,
    coerce=True,
)

ofgem_monthly_schema = pa.DataFrameSchema(
    {
        "year": pa.Column("int64", coerce=True),
        "month": pa.Column("int64", coerce=True, checks=pa.Check.in_range(1, 12)),
        "country": pa.Column(str, checks=pa.Check.isin(["GB", "NI"])),
        "technology": pa.Column(str, coerce=True),
        "generation_mwh": pa.Column(float, nullable=True, coerce=True),
        "rocs_issued": pa.Column(float, nullable=True, coerce=True),
    },
    strict=False,
    coerce=True,
)


# ---------------------------------------------------------------------------
# 12-year XLSX download — D-17 fail-loud (output_path bound BEFORE try).
# ---------------------------------------------------------------------------
def download_twelve_year_xlsx() -> Path:
    """Download the Ofgem 12-year RO aggregate XLSX (2006-2018) into ``data/raw/ofgem/``.

    Returns the path to the written file on success.

    Raises:
        requests.exceptions.RequestException: on any network failure
            (D-17 fail-loud — daily refresh workflow needs to see the failure
            propagate, not a silently un-downloaded path).
    """
    output_path = DATA_DIR / "raw" / "ofgem" / XLSX_FILENAME  # BOUND BEFORE try (D-17 gap #2 fix)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        response = requests.get(
            XLSX_URL, headers=HEADERS, stream=True, timeout=60
        )
        response.raise_for_status()
        with open(output_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        return output_path
    except requests.exceptions.RequestException as e:
        print(f"An error occurred while downloading 12-year XLSX: {e}")
        raise  # fail loud per D-17


# ---------------------------------------------------------------------------
# XLSX → monthly DataFrame — real implementation (Plan 02 replaces stub).
# ---------------------------------------------------------------------------
def parse_xlsx_to_monthly() -> pd.DataFrame:
    """Parse the 12-year Ofgem XLSX into a deterministic monthly aggregate.

    Return-shape contract: a DataFrame conforming to ``ofgem_monthly_schema``
    (columns: ``year``, ``month``, ``country``, ``technology``,
    ``generation_mwh``, ``rocs_issued``).

    Source sheet: "ROCs by Tech & Month" in the 12-year XLSX.  Structure:
      Row: "<YYYY/YYYY>" (year header, e.g. "2006/7" or "2007/2008")
      Row: "Technology" | "Apr" | "May" | ... | "Mar" | "TOTAL" (header)
      Rows: technology name + 12 monthly ROC counts + annual TOTAL
      Row: "TOTAL" + 12 monthly sums + grand total

    The sheet covers Apr 2006 – Mar 2018 (12 complete obligation years).
    No per-country split or generation MWh is available in this sheet;
    ``country`` is set to ``"GB"`` (UK aggregate) and ``generation_mwh``
    is ``None`` (null) throughout.

    Determinism discipline (Phase 4 D-21): same XLSX bytes → byte-identical
    output DataFrame across runs. No clock, no randomness; technology names
    sorted lexicographically within each year-month group; final sort on
    (year, month, country, technology).

    Returns
    -------
    pd.DataFrame
        Monthly aggregate validated against ``ofgem_monthly_schema``.
        Caller (``emit_ro_generation_csv``) writes this to
        ``data/raw/ofgem/ro-generation.csv`` with a provenance header.

    Raises
    ------
    FileNotFoundError
        If the 12-year XLSX is absent from ``data/raw/ofgem/``.
    ValueError
        If the sheet structure cannot be parsed (year-header not found or
        unexpected column count).
    """
    import openpyxl

    xlsx_path = DATA_DIR / "raw" / "ofgem" / XLSX_FILENAME
    if not xlsx_path.exists():
        raise FileNotFoundError(
            f"12-year XLSX not found — run download_twelve_year_xlsx() first: {xlsx_path}"
        )

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb["ROCs by Tech & Month"]

    # Month-name → calendar month number (Apr=4, ..., Mar=3 of next year).
    # The XLSX reports Apr–Mar within a single obligation year; we map to
    # calendar months so rows carry (year, month) pairs that match Apr=4 in
    # the start calendar year, through Mar=3 in the next calendar year.
    _MONTH_COLS: dict[str, int] = {
        "Apr": 4, "May": 5, "Jun": 6, "Jul": 7, "Aug": 8, "Sep": 9,
        "Oct": 10, "Nov": 11, "Dec": 12, "Jan": 1, "Feb": 2, "Mar": 3,
    }
    _MONTH_ORDER = ["Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar"]

    records: list[dict] = []
    current_start_year: int | None = None
    in_data_section = False

    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            in_data_section = False
            continue

        cell0 = row[0]
        if cell0 is None:
            continue

        cell0_str = str(cell0).strip()

        # Detect year-header rows like "2006/7" or "2007/2008".
        if "/" in cell0_str and len(cell0_str) <= 10 and cell0_str[0].isdigit():
            # Parse start year from e.g. "2006/7" → 2006
            try:
                current_start_year = int(cell0_str.split("/")[0])
                in_data_section = False  # wait for Technology header
            except ValueError:
                pass
            continue

        # Detect "Technology | Apr | May | ..." column-header row.
        if cell0_str == "Technology" and current_start_year is not None:
            in_data_section = True
            continue

        # Skip TOTAL aggregate rows and non-data rows.
        if cell0_str in {"TOTAL", "Return to information tab"} or not in_data_section:
            continue

        # Data row: technology name + 12 monthly counts.
        technology = cell0_str
        month_values = row[1:13]  # columns B through M (Apr..Mar)

        if current_start_year is None:
            continue

        for col_idx, month_name in enumerate(_MONTH_ORDER):
            raw_val = month_values[col_idx]
            rocs = None
            if raw_val is not None:
                try:
                    rocs = float(raw_val)
                except (TypeError, ValueError):
                    rocs = None

            # Map Apr–Dec to start year; Jan–Mar to start_year+1.
            cal_month = _MONTH_COLS[month_name]
            cal_year = current_start_year if cal_month >= 4 else current_start_year + 1

            records.append({
                "year": cal_year,
                "month": cal_month,
                "country": "GB",
                "technology": technology,
                "generation_mwh": None,
                "rocs_issued": rocs,
            })

    wb.close()

    if not records:
        raise ValueError(
            f"No data rows parsed from 'ROCs by Tech & Month' in {xlsx_path}. "
            "Verify XLSX structure matches expected year-header + Technology format."
        )

    df = pd.DataFrame(records)
    # Sort deterministically (D-21): by (year, month, country, technology).
    df = df.sort_values(
        ["year", "month", "country", "technology"],
        kind="mergesort",
    ).reset_index(drop=True)

    # Cast to schema dtypes before validation.
    df["year"] = df["year"].astype("int64")
    df["month"] = df["month"].astype("int64")
    df["rocs_issued"] = pd.to_numeric(df["rocs_issued"], errors="coerce")

    return ofgem_monthly_schema.validate(df)


def emit_ro_generation_csv(output_path: Path | None = None) -> Path:
    """Parse the 12-year XLSX and write ``ro-generation.csv`` with provenance header.

    This is the canonical entry point that persists the monthly aggregate to
    ``data/raw/ofgem/ro-generation.csv`` and updates its ``.meta.json`` sidecar.

    Determinism (D-21): ``parse_xlsx_to_monthly()`` → sort → write. No clock
    or randomness in the DataFrame; sidecar ``retrieved_at`` is non-deterministic
    (current UTC timestamp) but the CSV bytes themselves are byte-identical.

    Returns
    -------
    Path
        Path to the written CSV file.
    """
    from uk_subsidy_tracker.data.sidecar import _sha256_of, write_sidecar

    out = output_path or (DATA_DIR / "raw" / "ofgem" / "ro-generation.csv")

    df = parse_xlsx_to_monthly()

    provenance_lines = [
        "# Provenance: Monthly ROC-issue aggregate regenerated from Ofgem 12-year XLSX (D-01).",
        "# Source: data/raw/ofgem/rocs_report_2006_to_2018_20250410081520.xlsx",
        "#   Sheet: 'ROCs by Tech & Month'",
        "#   Coverage: Apr 2006 – Mar 2018 (12 obligation years SY5–SY17)",
        "# generation_mwh is null throughout — the 12-year XLSX carries ROC counts only.",
        "# country='GB' throughout — the sheet aggregates all UK nations.",
        "#",
        "# Re-generate: uv run python3 -c",
        "#   \"from uk_subsidy_tracker.data.ofgem_aggregate import emit_ro_generation_csv; emit_ro_generation_csv()\"",
        "#",
        f"# SHA256 of source XLSX: {_sha256_of(DATA_DIR / 'raw' / 'ofgem' / XLSX_FILENAME)}",
    ]

    out.parent.mkdir(parents=True, exist_ok=True)
    with out.open("w", encoding="utf-8", newline="") as f:
        for line in provenance_lines:
            f.write(line + "\n")
        df.to_csv(f, index=False)

    # Update sidecar to chain derivation back to the 12-year XLSX sha256.
    xlsx_path = DATA_DIR / "raw" / "ofgem" / XLSX_FILENAME
    write_sidecar(
        raw_path=out,
        upstream_url=XLSX_URL,
        http_status=None,
        sources=[
            {
                "url": XLSX_URL,
                "sha256": _sha256_of(xlsx_path),
                "retrieved_on": "2026-04-24",
                "notes": (
                    "Ofgem 12-year XLSX (Apr 2006 – Mar 2018); "
                    "ro-generation.csv derived from 'ROCs by Tech & Month' sheet via parse_xlsx_to_monthly()"
                ),
            }
        ],
    )
    return out


# ---------------------------------------------------------------------------
# Transcribed-CSV loaders — pandera-validated, comment-aware.
# ---------------------------------------------------------------------------
def load_annual_aggregate_csv() -> pd.DataFrame:
    """Load + validate the SY17–SY23 transcribed annual-aggregate CSV.

    The CSV file begins with a ``# Provenance:`` block (D-03) listing each
    source PDF URL + sha256 + retrieved_on. ``pandas.read_csv(comment="#")``
    skips those header lines so only data rows reach the pandera validator.
    """
    path = DATA_DIR / "raw" / "ofgem" / "ro-annual-aggregate.csv"
    df = pd.read_csv(path, comment="#")
    df = ofgem_annual_aggregate_schema.validate(df)
    return df


def load_roc_prices_csv() -> pd.DataFrame:
    """Load + validate the transcribed ROC-prices CSV (multi-source provenance).

    The CSV file begins with a ``# Provenance:`` block (D-03) listing each
    source PDF URL + sha256 + retrieved_on. ``pandas.read_csv(comment="#")``
    skips those header lines so only data rows reach the pandera validator.
    """
    path = DATA_DIR / "raw" / "ofgem" / "roc-prices.csv"
    df = pd.read_csv(path, comment="#")
    df = ofgem_roc_prices_schema.validate(df)
    return df


__all__ = [
    "XLSX_URL",
    "XLSX_FILENAME",
    "ofgem_annual_aggregate_schema",
    "ofgem_roc_prices_schema",
    "ofgem_monthly_schema",
    "download_twelve_year_xlsx",
    "parse_xlsx_to_monthly",
    "emit_ro_generation_csv",
    "load_annual_aggregate_csv",
    "load_roc_prices_csv",
    "OfgemAnnualReportConfig",
    "OfgemAnnualReportsConfig",
    "load_ofgem_annual_reports_config",
    "download_annual_xlsx",
    "parse_annual_xlsx_to_aggregate_rows",
    "emit_annual_aggregate_csv",
]


# ----- Phase 05.2 Plan 02 (revised): SY18-SY23 annual-report XLSX path -----


class OfgemAnnualReportConfig(BaseModel):
    scheme_year: str
    period: str
    url: str
    local_filename: str
    expected_min_size_bytes: int
    notes: str


class OfgemAnnualReportsConfig(BaseModel):
    reports: list[OfgemAnnualReportConfig]

    def by_scheme_year(self, scheme_year: str) -> OfgemAnnualReportConfig:
        for r in self.reports:
            if r.scheme_year == scheme_year:
                return r
        raise KeyError(scheme_year)


_ANNUAL_REPORTS_YAML = Path(__file__).parent / "ofgem_annual_reports.yaml"


def load_ofgem_annual_reports_config(path: Path | None = None) -> OfgemAnnualReportsConfig:
    """Load the SY18-SY23 XLSX-companion manifest.

    Provenance: data/ofgem_annual_reports.yaml carries the verbatim URL inventory
    from CONTEXT.md (SY17-SY23 discovery, 2026-04-24, commit b934492).
    """
    target = path if path is not None else _ANNUAL_REPORTS_YAML
    with open(target, "r", encoding="utf-8") as f:
        raw = yaml.safe_load(f)
    return OfgemAnnualReportsConfig(**raw)


def download_annual_xlsx(scheme_year: str) -> Path:
    """Download a single SY18-SY23 annual-report XLSX dataset companion.

    Resolves URL via the YAML manifest. D-17 fail-loud discipline:
      - output_path bound BEFORE try (gap #2 fix from Phase 4 Plan 07)
      - timeout=60 on requests.get
      - bare raise in except RequestException

    Raises KeyError if scheme_year not in manifest (e.g. 'SY17' deferred).
    """
    cfg = load_ofgem_annual_reports_config()
    entry = cfg.by_scheme_year(scheme_year)  # raises KeyError on SY17
    output_path = DATA_DIR / "raw" / "ofgem" / entry.local_filename  # BOUND BEFORE try (D-17 gap #2)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        response = requests.get(entry.url, headers=HEADERS, stream=True, timeout=60)
        response.raise_for_status()
        with open(output_path, "wb") as f:
            for chunk in response.iter_content(chunk_size=8192):
                f.write(chunk)
        return output_path
    except requests.exceptions.RequestException as e:
        print(f"An error occurred while downloading {scheme_year} XLSX: {e}")
        raise  # fail loud per D-17


def _scheme_year_end_calendar(scheme_year: str) -> int:
    """Return the end calendar year for a scheme_year label.

    Scheme years run April-March:
      SY18 = 2019-20 → end calendar year 2020
      SY19 = 2020-21 → end calendar year 2021
      SY20 = 2021-22 → end calendar year 2022
      SY21 = 2022-23 → end calendar year 2023
      SY22 = 2023-24 → end calendar year 2024
      SY23 = 2024-25 → end calendar year 2025
    """
    _MAP: dict[str, int] = {
        "SY18": 2020, "SY19": 2021, "SY20": 2022,
        "SY21": 2023, "SY22": 2024, "SY23": 2025,
    }
    if scheme_year not in _MAP:
        raise KeyError(f"Unknown scheme_year for calendar mapping: {scheme_year!r}")
    return _MAP[scheme_year]


def _clean_roc_value(v: object) -> float | None:
    """Convert a cell value to float ROC count. '-' and None become None."""
    if v is None or v == "-" or str(v).strip() == "-":
        return None
    try:
        return float(str(v).replace(",", ""))
    except (ValueError, TypeError):
        return None


def _parse_sy18(wb: object, entry_url: str, cal_year: int) -> list[dict]:
    """SY18 (2019-20): Only aggregate totals available — no per-technology breakdown.

    Sheet: '2. ROCs issued and generation'
    Row 7 carries 'Associated renewable generation (MWh)' = 84,920,897 (UK total).
    Row 6 carries 'Total number of ROCs issued' = 114,706,958 (UK total).
    No per-technology split; emits one row with technology='All technologies', country='GB'.

    This is the sole year for which technology disaggregation is not available
    in the published XLSX dataset companion.
    """
    import openpyxl

    ws = wb["2. ROCs issued and generation"]
    rocs_total = None
    gen_mwh_total = None
    for row in ws.iter_rows(values_only=True):
        if row[0] is not None:
            cell0 = str(row[0])
            if "Total number of ROCs issued" in cell0:
                rocs_total = _clean_roc_value(row[1])
            elif "Associated renewable generation" in cell0 and "MWh" in cell0:
                gen_mwh_total = _clean_roc_value(row[1])
    gen_gwh = (gen_mwh_total / 1000) if gen_mwh_total is not None else None
    return [{
        "scheme_year": "SY18",
        "year": cal_year,
        "country": "GB",
        "technology": "All technologies",
        "generation_gwh": gen_gwh,
        "rocs_issued": rocs_total,
        "ro_cost_gbp_nominal": None,
        "source_pdf_url": entry_url,
    }]


def _parse_sy19(wb: object, entry_url: str, cal_year: int) -> list[dict]:
    """SY19 (2020-21): Technology × total-UK ROC data from 'Table 2.2'.

    Header row: ['Technology', 'England', 'Scotland', 'Wales', 'Northern Ireland', 'Total']
    Data rows: technology name in col B, UK total ROCs in col F.
    No per-technology generation MWh available; generation_gwh set to None.
    Country = 'GB' (UK total; NI included in ROC total for this year).
    """
    ws = wb["Table 2.2"]
    rows: list[dict] = []
    skip_terms = {"Total", "Return to information tab"}
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[1] is None:
            continue
        cell1 = str(row[1]).strip()
        if cell1 == "Technology":
            header_found = True
            continue
        if not header_found:
            continue
        if cell1 in skip_terms or cell1 == "":
            continue
        # col F (index 6) = Total
        rocs_total = _clean_roc_value(row[6])
        rows.append({
            "scheme_year": "SY19",
            "year": cal_year,
            "country": "GB",
            "technology": cell1,
            "generation_gwh": None,
            "rocs_issued": rocs_total,
            "ro_cost_gbp_nominal": None,
            "source_pdf_url": entry_url,
        })
    return rows


def _parse_sy20_sy23(wb: object, entry_url: str, scheme_year: str, cal_year: int) -> list[dict]:
    """SY20-SY23: Technology × Country from 'Figure 3.2' (ROCs) + 'Figure 3.3' (MWh).

    Sheet 'Figure 3.2': header at row 8 (0-indexed):
      ['Technology', 'England', 'Scotland', 'Wales', 'Northern Ireland', 'Total']
    Sheet 'Figure 3.3': same header, 'Total (MWh)' in last data col.

    Strategy: emit one row per (technology, country) using per-country values.
    Countries: England, Scotland, Wales, Northern Ireland.
    '-' values are emitted as None (technology not active in that country).
    """
    countries = ["England", "Scotland", "Wales", "Northern Ireland"]
    country_code = {"England": "GB", "Scotland": "GB", "Wales": "GB", "Northern Ireland": "NI"}

    # --- ROCs (Figure 3.2) ---
    ws_roc = wb["Figure 3.2"]
    skip_terms = {"Total", "Return to information tab", ""}
    roc_by_tech_country: dict[tuple[str, str], float | None] = {}
    header_found = False
    for row in ws_roc.iter_rows(values_only=True):
        if row[1] is None:
            continue
        cell1 = str(row[1]).strip()
        if cell1 == "Technology":
            header_found = True
            continue
        if not header_found:
            continue
        if cell1 in skip_terms:
            continue
        # cols 2-5 = England, Scotland, Wales, NI
        for ci, country in enumerate(countries):
            roc_val = _clean_roc_value(row[ci + 2])
            roc_by_tech_country[(cell1, country)] = roc_val

    # --- Generation MWh (Figure 3.3) ---
    ws_gen = wb["Figure 3.3"]
    gen_by_tech_country: dict[tuple[str, str], float | None] = {}
    header_found = False
    for row in ws_gen.iter_rows(values_only=True):
        if row[1] is None:
            continue
        cell1 = str(row[1]).strip()
        if cell1 == "Technology":
            header_found = True
            continue
        if not header_found:
            continue
        if cell1 in skip_terms or "Total (MWh)" in cell1:
            continue
        for ci, country in enumerate(countries):
            gen_mwh = _clean_roc_value(row[ci + 2])
            gen_gwh = (gen_mwh / 1000) if gen_mwh is not None else None
            gen_by_tech_country[(cell1, country)] = gen_gwh

    # --- Emit rows ---
    rows: list[dict] = []
    all_techs = sorted({k[0] for k in roc_by_tech_country})
    for tech in all_techs:
        for country in countries:
            rocs = roc_by_tech_country.get((tech, country))
            gen_gwh = gen_by_tech_country.get((tech, country))
            # Skip rows where both ROCs and generation are None (technology absent)
            if rocs is None and gen_gwh is None:
                continue
            rows.append({
                "scheme_year": scheme_year,
                "year": cal_year,
                "country": country_code[country],
                "technology": tech,
                "generation_gwh": gen_gwh,
                "rocs_issued": rocs,
                "ro_cost_gbp_nominal": None,
                "source_pdf_url": entry_url,
            })
    return rows


def parse_annual_xlsx_to_aggregate_rows(scheme_year: str) -> pd.DataFrame:
    """Parse one SY18-SY23 XLSX into ofgem_annual_aggregate_schema rows.

    Returns DataFrame with columns: scheme_year, year, country, technology,
    generation_gwh, rocs_issued, ro_cost_gbp_nominal, source_pdf_url.
    Under the XLSX path: source_pdf_url carries the XLSX URL.

    Sheet inventory (determined via Step 1 inspection, hard-coded):
      SY18: '2. ROCs issued and generation' — aggregate totals only; technology='All technologies'
      SY19: 'Table 2.2' — technology × UK-total ROCs; no per-country MWh breakdown
      SY20-SY23: 'Figure 3.2' (ROCs by tech × country) + 'Figure 3.3' (MWh by tech × country)

    Pure function: same XLSX bytes → byte-identical DataFrame (D-21).
    """
    import openpyxl

    cfg = load_ofgem_annual_reports_config()
    entry = cfg.by_scheme_year(scheme_year)
    xlsx_path = DATA_DIR / "raw" / "ofgem" / entry.local_filename
    cal_year = _scheme_year_end_calendar(scheme_year)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if scheme_year == "SY18":
            records = _parse_sy18(wb, entry.url, cal_year)
        elif scheme_year == "SY19":
            records = _parse_sy19(wb, entry.url, cal_year)
        else:
            records = _parse_sy20_sy23(wb, entry.url, scheme_year, cal_year)
    finally:
        wb.close()

    df = pd.DataFrame(records, columns=[
        "scheme_year", "year", "country", "technology",
        "generation_gwh", "rocs_issued", "ro_cost_gbp_nominal", "source_pdf_url",
    ])
    df = df.sort_values(["country", "technology"], kind="mergesort").reset_index(drop=True)
    df = ofgem_annual_aggregate_schema.validate(df)
    return df


def emit_annual_aggregate_csv(output_path: Path | None = None) -> Path:
    """Concatenate SY18-SY23 parsed rows into ro-annual-aggregate.csv.

    Writes:
      1. # Provenance: comment header listing the 6 source XLSX URLs.
      2. Column header + data rows.
      3. Sibling .meta.json sidecar with sources[] enumerating each XLSX's
         url + sha256 + retrieved_on + scheme_year + notes.
    """
    import hashlib
    from datetime import date
    from uk_subsidy_tracker.data.sidecar import write_sidecar

    target = output_path if output_path is not None else (
        DATA_DIR / "raw" / "ofgem" / "ro-annual-aggregate.csv"
    )

    cfg = load_ofgem_annual_reports_config()
    frames = [parse_annual_xlsx_to_aggregate_rows(r.scheme_year) for r in cfg.reports]
    unified = pd.concat(frames, ignore_index=True)
    unified = unified.sort_values(
        ["scheme_year", "country", "technology"], kind="mergesort"
    ).reset_index(drop=True)

    # Build the # Provenance: header block
    retrieved_on = date.today().isoformat()
    header_lines = [
        "# Provenance: RO annual-aggregate emitted from Ofgem XLSX dataset companions (SY18-SY23).",
        "#",
        "# Each row below is extracted from the scheme-year XLSX listed in the .meta.json sources[] array.",
        "# SY17 deferred per Phase 05.2 revision_decisions.sy17_disposition (PDF-only year).",
        "#",
        "# Source XLSX inventory (also enumerated in .meta.json):",
    ]
    for r in cfg.reports:
        header_lines.append(f"#   - {r.scheme_year} ({r.period}): {r.url}")
    header_lines.append("#")
    header_lines.append("# Audit log:")
    header_lines.append(
        f"#   {retrieved_on} — Phase 05.2 Plan 02 emission (replaces commit a720ae7 in-place)."
    )
    header_lines.append("#")

    with open(target, "w", encoding="utf-8") as f:
        f.write("\n".join(header_lines) + "\n")
        unified.to_csv(f, index=False, lineterminator="\n")

    # Build sources[] for sidecar
    sources = []
    for r in cfg.reports:
        xlsx_path = DATA_DIR / "raw" / "ofgem" / r.local_filename
        sha256 = hashlib.sha256(xlsx_path.read_bytes()).hexdigest()
        sources.append({
            "url": r.url,
            "sha256": sha256,
            "retrieved_on": retrieved_on,
            "scheme_year": r.scheme_year,
            "notes": r.notes,
        })

    write_sidecar(
        raw_path=target,
        upstream_url="emitted-from-ofgem-annual-xlsxes",
        http_status=None,
        publisher_last_modified=None,
        sources=sources,
    )
    return target
